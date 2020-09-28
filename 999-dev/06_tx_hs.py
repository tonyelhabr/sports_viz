#%%
import openpyxl
import os
#%%
path = '../../data-raw/0x/Texas High School Football Team Public Links.xlsm'
os.path.exists(path)
#%%
wb = openpyxl.load_workbook(path)
type(wb)
#%%
print(wb.sheetnames)
#%%
df = wb['Data']
#%%
c = df['url'][0]
#%%
df.cell(2, 5).value
#%%
df.max_row
#%%
import requests
url = df.cell(2, 5).value
resp = requests.get(url)
res = open('test.xlsx', 'wb')
res.write(resp.content)
res.close()
#%%
